from PyQt6.QtWidgets import QDialog, QLineEdit, QFormLayout, QPushButton
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import datetime as dt
import inspect

# Path to the Excel tracker file.
# NOTE: Replace {Username} with an actual user or load dynamically from config.
tracker_path = r"C:\Users\{Username}\Path_to_tracker"

# ----------------------------------------------------------
# GENERAL NOTES ABOUT THIS MODULE
# ----------------------------------------------------------
# - This script updates multiple Excel-based trackers using openpyxl.
# - It assumes that specific tables exist inside specific sheets.
# - Column indexes are hardcoded (e.g., col_to_check = 4), so if the Excel
#   structure changes, these indices must be updated.
# - openpyxl stores Excel tables in a private attribute `_tables`.
#   This works, but is not part of the official stable API.
# - Several functions compare cell values to input text. These comparisons
#   may fail if types differ (e.g., comparing '12' to 12). Comments below
#   highlight where this is relevant.
# ----------------------------------------------------------


def update_tracker_interview(window):
    """
    Appends a new interview entry to the 'Interviews' sheet.
    Expected table columns:
        ID, Interview Date, Candidate, Interviewers, Stage, Awaiting acceptance, Case Study
    """
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["Interviews"]

    # Retrieve Excel table from sheet (first table found).
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return

    table = tables[0]

    # Table reference looks like "A1:G10". Split into start and end cell.
    start_cell, end_cell = table.ref.split(":")
    start_row = int("".join(filter(str.isdigit, start_cell)))
    start_col = "".join(filter(str.isalpha, start_cell))
    end_col = "".join(filter(str.isalpha, end_cell))
    end_row = int("".join(filter(str.isdigit, end_cell)))

    # Place the new entry directly under the last table row.
    first_empty_row = end_row + 1

    # NOTE: window.id must be convertible to int; otherwise this will error.
    values = [int(window.id), None, window.name, None, "Placeholder", "Candidate"]

    # Write data into the new row and center-align text.
    for col_index, value in enumerate(values, start=1):
        cell = tracker_tab.cell(row=first_empty_row, column=col_index, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Extend the table range to include the new row.
    # WARNING: if table has filters/styles, this may not carry formatting downward.
    table.ref = f"{start_col}{start_row}:{end_col}{first_empty_row}"

    tracker_workbook.save(tracker_path)

def update_tracker_interview_case_study(window):
    """
    Updates the 'Case Study' column inside the 'Interviews' table.
    Match is done using candidate ID.
    """
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["Interviews"]

    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return

    table = tables[0]

    # Skip header row by adding +1
    start_cell, end_cell = table.ref.split(":")
    start_row = int("".join(filter(str.isdigit, start_cell))) + 1
    end_row = int("".join(filter(str.isdigit, end_cell)))
    col_to_check = 1  # Candidate ID column
    col_to_update = 8  # Case study update column

    for row in range(start_row, end_row + 1):
        cell = tracker_tab.cell(row=row, column=col_to_check)

        # Handle ID type differences defensively.
        try:
            target_id = int(window.id)
        except:
            # Fallback: compare as strings if the ID is not numeric.
            if str(cell.value).strip() == str(window.id).strip():
                tracker_tab.cell(
                    row=row,
                    column=col_to_update,
                    value=f"case study sent {dt.date.today():%d.%m}",
                )
            continue

        # If IDs match, update the cell value.
        if cell.value == target_id:
            tracker_tab.cell(
                row=row,
                column=col_to_update,
                value=f"case study sent {dt.date.today():%d.%m}",
            )

    tracker_workbook.save(tracker_path)

def update_tracker_new_hire_CN_chaser(window):
    # Load the Excel workbook and select the "New hire tracker" sheet
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["New hire tracker"]

    # Retrieve all Excel tables defined inside the sheet
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return
    table = tables[0]
    
    # Extract the table range (e.g. "A1:H20") and determine row boundaries
    start_cell, end_cell = table.ref.split(":")
    start_row = int(''.join(filter(str.isdigit, start_cell))) + 1  # Skip header row
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    # Column 4 = Candidate name; Column 6 = Onboarding task notes
    col_to_check = 4
    col_to_update = 6

    # Loop through each row inside the table
    for row in range(start_row, end_row + 1):
        col_check_cell = tracker_tab.cell(row=row, column=col_to_check)
        col_update_cell = tracker_tab.cell(row=row, column=col_to_update)

        # Check if this row belongs to the candidate entered in the form
        if col_check_cell.value == window.name:
            # Retrieve existing cell value, or empty string if cell is blank
            current_value = col_update_cell.value or ""

            # Logic for updating the text depending on what is already stored
            if current_value == "HM&CN tasks TBC":
                col_update_cell.value = f"HM tasks TBC, CN chaser {dt.date.today():%d.%m}"
            elif current_value == "CN tasks TBC":
                col_update_cell.value = f"CN chaser {dt.date.today():%d.%m}"
            else:
                # Append the new chaser info to any existing notes
                col_update_cell.value = current_value + f", CN chaser {dt.date.today():%d.%m}"
            
            # Apply cell alignment so text is centered
            col_update_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save changes back to the tracker file
    tracker_workbook.save(tracker_path)

def update_tracker_iForm_chaser(window):
    # Load workbook and sheet
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["New hire tracker"]

    # Ensure the sheet contains an Excel table
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return
    table = tables[0]
    
    # Determine row limits for the table
    start_cell, end_cell = table.ref.split(":")
    start_row = int(''.join(filter(str.isdigit, start_cell))) + 1
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    # Columns used for searching/updating
    col_to_check = 4
    col_to_update = 6

    # Loop through rows and update the correct one
    for row in range(start_row, end_row + 1):
        cell = tracker_tab.cell(row=row, column=col_to_check)
        if cell.value == window.name:
            # Write an "iForm chaser" note with today's date
            tracker_tab.cell(
                row=row,
                column=col_to_update,
                value="".join(['iForm chaser ', dt.date.today().strftime("%d.%m")])
            )

    tracker_workbook.save(tracker_path)

def update_tracker_NIN_chaser(window):
    # Load workbook and sheet
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["New hire tracker"]

    # Retrieve table metadata
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return
    table = tables[0]
    
    # Determine table range
    start_cell, end_cell = table.ref.split(":")
    start_row = int(''.join(filter(str.isdigit, start_cell))) + 1
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    # Candidate name and update columns
    col_to_check = 4
    col_to_update = 6

    for row in range(start_row, end_row + 1):
        cell = tracker_tab.cell(row=row, column=col_to_check)

        if cell.value == window.name:
            # Insert NIN chaser update
            tracker_tab.cell(
                row=row,
                column=col_to_update,
                value="".join(['NIN chaser ', dt.date.today().strftime("%d.%m")])
            )

    tracker_workbook.save(tracker_path)

def update_tracker_new_hire_HM_chaser(window):
    # Load workbook and sheet
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["New hire tracker"]

    # Retrieve Excel table
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return
    table = tables[0]
    
    # Parse table boundaries
    start_cell, end_cell = table.ref.split(":")
    start_row = int(''.join(filter(str.isdigit, start_cell))) + 1
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    col_to_check = 4
    col_to_update = 6

    for row in range(start_row, end_row + 1):
        col_check_cell = tracker_tab.cell(row=row, column=col_to_check)
        col_update_cell = tracker_tab.cell(row=row, column=col_to_update)

        # Match candidate name from form input
        if col_check_cell.value == window.CN_name:
            current_value = col_update_cell.value or ""

            # Apply update logic depending on existing notes
            if current_value == "HM&CN tasks TBC":
                col_update_cell.value = f"CN tasks TBC, HM chaser {dt.date.today():%d.%m}"
            elif current_value == "HM tasks TBC":
                col_update_cell.value = f"HM chaser {dt.date.today():%d.%m}"
            else:
                col_update_cell.value = current_value + f", HM chaser {dt.date.today():%d.%m}"

            col_update_cell.alignment = Alignment(horizontal='center', vertical='center')

    tracker_workbook.save(tracker_path)

def update_tracker_policies_chaser(window):
    # Load workbook and sheet
    tracker_workbook = load_workbook(tracker_path)
    tracker_tab = tracker_workbook["New hire tracker"]

    # Retrieve table metadata
    tables = list(tracker_tab._tables.values())
    if not tables:
        print("Couldn't find any table in workbook!")
        return
    table = tables[0]
    
    # Extract row boundaries
    start_cell, end_cell = table.ref.split(":")
    start_row = int(''.join(filter(str.isdigit, start_cell))) + 1
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    col_to_check = 4
    col_to_update = 6

    for row in range(start_row, end_row + 1):
        cell = tracker_tab.cell(row=row, column=col_to_check)

        if cell.value == window.CN_name:
            # Record policy chaser update
            tracker_tab.cell(
                row=row,
                column=col_to_update,
                value="".join(['Onboarding policies chaser ', dt.date.today().strftime("%d.%m")])
            )

    tracker_workbook.save(tracker_path)

class basic_form(QDialog):
    # Capture the function and module that created this form.
    # Used to determine what action should happen after clicking "Proceed".
    stack = inspect.stack()
    caller_frame = stack[1]
    caller_function_name = caller_frame.function
    caller_module = inspect.getmodule(caller_frame.frame).__name__

    def __init__(self, caller_module):
        super().__init__()
        self.caller_module = caller_module
        
        self.setWindowTitle("Form")

        # Basic input fields for candidate information
        self.name_input = QLineEdit()
        self.email_input = QLineEdit()

        # Layout containing labels and input widgets
        layout = QFormLayout()
        layout.addRow("Candidate's name:", self.name_input)
        layout.addRow("Candidate's email:", self.email_input)

        # Button to confirm the form and trigger the update logic
        self.save_button = QPushButton("Proceed")
        self.save_button.clicked.connect(self.save_data)

        layout.addRow(self.save_button)

        self.setLayout(layout)

    def save_data(self):
        # Store entered data on the object so tracker functions can access it
        self.name = self.name_input.text()
        self.email = self.email_input.text()

        # Determine which tracker update function to run based on module name
        if self.caller_module == "Onboarding_CN_chaser":
            update_tracker_new_hire_CN_chaser(self)
        elif self.caller_module == "iForm_chaser":
            update_tracker_iForm_chaser(self)
        elif self.caller_module == "NIN_chaser":
            update_tracker_NIN_chaser(self)

        # Close the form after processing
        self.close()

class CN_availability(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Candidate's availability request form")

        # Input fields used for interview scheduling / availability
        self.id_input = QLineEdit()
        self.job_title_input = QLineEdit()
        self.name_input = QLineEdit()
        self.email_input = QLineEdit()
        self.duration_input = QLineEdit()
        self.iv_format_input = QLineEdit()
        self.address_input = QLineEdit()
        self.slot_input = QLineEdit()

        # Form layout with labels
        layout = QFormLayout()
        layout.addRow("ID:", self.id_input)
        layout.addRow("Job title:", self.job_title_input)
        layout.addRow("Candidate's name:", self.name_input)
        layout.addRow("Candidate's email:", self.email_input)
        layout.addRow("Duration:", self.duration_input)
        layout.addRow("Format:", self.iv_format_input)
        layout.addRow("Address:", self.address_input)
        layout.addRow("Slot:", self.slot_input)

        # Confirmation button
        self.save_button = QPushButton("Proceed")
        self.save_button.clicked.connect(self.save_data)

        layout.addRow(self.save_button)

        self.setLayout(layout)

    def save_data(self):
        # Store all inputs for downstream functions
        self.id = self.id_input.text()
        self.job_title = self.job_title_input.text()
        self.name = self.name_input.text()
        self.email = self.email_input.text()
        self.duration = self.duration_input.text()
        self.ivformat = self.iv_format_input.text()
        self.address = self.address_input.text()
        self.slot = self.slot_input.text()

        # Update tracker for interview / availability
        update_tracker_interview(self)
        
        self.close()

class HM_HRBP_chaser(QDialog):
    def __init__(self, caller_module):
        super().__init__()
        self.caller_module = caller_module

        self.setWindowTitle("Onboarding Policies chaser form")

        # Input fields for hiring manager / candidate / HRBP information
        self.HM_name_input = QLineEdit()
        self.HM_email_input = QLineEdit()
        self.CN_name_input = QLineEdit()
        self.start_date_input = QLineEdit()
        self.HRBP_email_input = QLineEdit()

        # Form layout
        layout = QFormLayout()
        layout.addRow("Hiring Manager's name:", self.HM_name_input)
        layout.addRow("Hiring Manager's email:", self.HM_email_input)
        layout.addRow("Candidate's name:", self.CN_name_input)
        layout.addRow("Start date:", self.start_date_input)
        layout.addRow("HRBP's email:", self.HRBP_email_input)

        # Proceed button
        self.save_button = QPushButton("Proceed")
        self.save_button.clicked.connect(self.save_data)

        layout.addRow(self.save_button)

        self.setLayout(layout)

    def save_data(self):
        # Collect values from the form
        self.HM_name = self.HM_name_input.text()
        self.HM_email = self.HM_email_input.text()
        self.CN_name = self.CN_name_input.text()
        self.start_date = self.start_date_input.text()
        self.HRBP_email = self.HRBP_email_input.text()

        # Determine which onboarding action to execute
        if self.caller_module == "Onboarding_HM_chaser":
            update_tracker_new_hire_HM_chaser(self)
        elif self.caller_module == "policies_chaser":
            update_tracker_policies_chaser(self)
        
        self.close()


class case_study(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Candidate's availability request form")

        # Fields required for case study tracking
        self.id_input = QLineEdit()
        self.job_title_input = QLineEdit()
        self.name_input = QLineEdit()
        self.email_input = QLineEdit()

        layout = QFormLayout()
        layout.addRow("ID:", self.id_input)
        layout.addRow("Job title:", self.job_title_input)
        layout.addRow("Candidate's name:", self.name_input)
        layout.addRow("Candidate's email:", self.email_input)

        # Button to save and continue
        self.save_button = QPushButton("Proceed")
        self.save_button.clicked.connect(self.save_data)

        layout.addRow(self.save_button)

        self.setLayout(layout)

    def save_data(self):
        # Store entered case-study related data
        self.id = self.id_input.text()
        self.job_title = self.job_title_input.text()
        self.name = self.name_input.text()
        self.email = self.email_input.text()

        # Update tracker with case-study event
        update_tracker_interview_case_study(self)

        self.close()